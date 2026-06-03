"""
Análisis espacial y geográfico de vulnerabilidad por sequía y presión hídrica en Sonora.
Genera mapas detallados municipales y por DDR usando GeoPandas.
"""

import json
import os
from pathlib import Path

import geopandas as gpd
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns

# Configuración de rutas
BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
RAW_SHP = BASE_DIR / "data" / "raw" / "datos-sequia" / "impacto_sequia.shp"
INDICES_CSV = BASE_DIR / "data" / "processed" / "sequia_indices_sonora.csv"
REPNA1_CSV = BASE_DIR / "data" / "raw" / "datos_proporcionados" / "reporte-repna-1.csv"
REPNA2_CSV = BASE_DIR / "data" / "raw" / "datos_proporcionados" / "reporte-repna-2.csv"
CODIF_JSON = BASE_DIR / "data" / "config" / "codificacion.json"
OUTPUT_DIR = BASE_DIR / "reports" / "fase3" / "prueba_analisis" / "images"
REPORTS_DIR = BASE_DIR / "reports"
MONITOR_CSV = BASE_DIR / "data" / "processed" / "monitor_sequia_sonora.csv"
MAESTRO_CSV = (
    BASE_DIR / "data" / "processed" / "analisis_municipal_sonora_2010_2024.csv"
)

# Crear directorio si no existe
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Configurar estilo visual
sns.set_theme(style="whitegrid")
plt.rcParams.update(
    {
        "font.family": "sans-serif",
        "figure.titlesize": 16,
        "axes.titlesize": 14,
        "axes.labelsize": 11,
        "xtick.labelsize": 10,
        "ytick.labelsize": 10,
    }
)


def load_data():
    """Carga y limpia todos los conjuntos de datos requeridos."""
    # 1. Shapefile
    gdf = gpd.read_file(RAW_SHP)
    gdf_sonora = gdf[gdf["CVE_ENT"] == "26"].copy()
    gdf_sonora["CVE_MUN"] = gdf_sonora["CVE_MUN"].astype(int)

    # 2. Índices de sequía
    df_indices = pd.read_csv(INDICES_CSV)
    df_indices["CVE_MUN"] = df_indices["CVE_MUN"].astype(int)

    # 3. REPNA
    repna1 = pd.read_csv(REPNA1_CSV)
    repna2 = pd.read_csv(REPNA2_CSV)
    repna1 = repna1.rename(
        columns={"Volumen de extracción de aguas nacionales": "Volumen"}
    )
    repna2 = repna2.rename(
        columns={"Volumen de extracción de aguas nacionales (m3/año)": "Volumen"}
    )
    repna = pd.concat(
        [repna1[["Titular", "Volumen"]], repna2[["Titular", "Volumen"]]],
        ignore_index=True,
    )
    repna["Volumen"] = (
        repna["Volumen"].astype(str).str.replace(" ", "").str.replace(",", "")
    )
    repna["Volumen"] = pd.to_numeric(repna["Volumen"], errors="coerce").fillna(0)

    return gdf_sonora, df_indices, repna


def map_municipal_isag(gdf, df_indices):
    """Genera mapas vectoriales del índice ISAG municipal para ambos ciclos."""
    merged = gdf.merge(df_indices, on="CVE_MUN", how="inner")
    color_map = {1: "#d9534f", 2: "#f0ad4e", 3: "#5cb85c"}
    labels_map = {1: "Alta", 2: "Media", 3: "Baja"}

    for ciclo, col_isag, _ in [
        ("Primavera-Verano (P-V)", "ISAG_PV", "ISAG_PV_label"),
        ("Otoño-Invierno (O-I)", "ISAG_OI", "ISAG_OI_label"),
    ]:
        fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)

        for isag_val, color in color_map.items():
            subset = merged[merged[col_isag] == isag_val]
            if not subset.empty:
                subset.plot(
                    ax=ax,
                    color=color,
                    edgecolor="black",
                    linewidth=0.4,
                    label=labels_map[isag_val],
                )

        from matplotlib.patches import Patch

        legend_elements = [
            Patch(facecolor=color, edgecolor="black", label=labels_map[isag_val])
            for isag_val, color in color_map.items()
        ]

        ax.set_title(
            f"Índice de Impacto por Sequía Acumulada (ISAG)\nCiclo {ciclo} - Sonora",
            pad=15,
            weight="bold",
        )
        ax.axis("off")
        ax.legend(
            handles=legend_elements,
            title="Vulnerabilidad ISAG",
            loc="lower left",
            frameon=True,
            shadow=True,
        )

        filename = f"mapa_isag_{ciclo.split(' ')[0].lower()}.png"
        filepath = OUTPUT_DIR / filename
        plt.savefig(filepath, bbox_inches="tight")
        plt.close()


def analyze_ddr_pressure(gdf, df_indices, repna):
    """Calcula y grafica la presión hídrica de extracción por DDR."""
    siap_2024_path = (
        BASE_DIR
        / "data"
        / "processed"
        / "siap_produccion"
        / "sonora"
        / "cierre_agricola_sonora_2024.csv"
    )
    if siap_2024_path.exists():
        siap = pd.read_csv(siap_2024_path)
        mun_ddr_map = siap[["Nommunicipio", "Idddr"]].drop_duplicates()
        mun_ddr_map["Nommunicipio"] = mun_ddr_map["Nommunicipio"].str.strip()

        with open(CODIF_JSON, encoding="utf-8") as f:
            codif = json.load(f)
        ddr_names = codif["codigos_ddrs"]
        mun_ddr_map["DDR"] = mun_ddr_map["Idddr"].astype(str).map(ddr_names)

        df_merged = df_indices.merge(
            mun_ddr_map, left_on="Municipio", right_on="Nommunicipio", how="inner"
        )

        ddr_stats = (
            df_merged.groupby("DDR")
            .agg(
                {"REC_PV": "mean", "REC_OI": "mean", "CON_PV": "mean", "CON_OI": "mean"}
            )
            .reset_index()
        )

        fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
        ddr_stats_melted = ddr_stats.melt(
            id_vars="DDR",
            value_vars=["REC_PV", "REC_OI"],
            var_name="Ciclo",
            value_name="Recurrencia",
        )
        ddr_stats_melted["Ciclo"] = ddr_stats_melted["Ciclo"].map(
            {"REC_PV": "Primavera-Verano", "REC_OI": "Otoño-Invierno"}
        )

        sns.barplot(
            data=ddr_stats_melted,
            x="DDR",
            y="Recurrencia",
            hue="Ciclo",
            ax=ax,
            palette="Oranges_r",
        )
        ax.set_title(
            "Recurrencia Media de Sequía (%) por Distrito de Desarrollo Rural (DDR)",
            pad=15,
            weight="bold",
        )
        ax.set_ylabel("Recurrencia Promedio (%)")
        ax.set_xlabel("Distritos de Desarrollo Rural (DDR)")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()

        filepath = OUTPUT_DIR / "recurrencia_sequia_ddr.png"
        plt.savefig(filepath)
        plt.close()


def run_extended_analysis(gdf, df_indices):
    """Ejecuta los análisis cuantitativos faltantes (Frecuencia, Déficit, Concesiones)."""
    # 1. Frecuencia Estatal Histórica (2003-2026)
    df_mon = pd.read_csv(MONITOR_CSV)
    total_obs = len(df_mon)
    seq_obs = len(df_mon[df_mon["Categoria"].isin(["D1", "D2", "D3", "D4"])])
    pct_frec = (seq_obs / total_obs) * 100

    # 2. Déficit Hídrico Climatológico Promedio (ETo - Pef) por Municipio
    df_maestro = pd.read_csv(MAESTRO_CSV)
    df_deficit = (
        df_maestro.groupby("Municipio")
        .agg(
            {
                "ET0_total_mm": "mean",
                "Pef_total_mm": "mean",
                "deficit_hidrico_mm": "mean",
            }
        )
        .reset_index()
        .sort_values("deficit_hidrico_mm", ascending=False)
    )

    # Graficar municipios con mayor déficit climatológico
    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    sns.barplot(
        data=df_deficit.head(15),
        x="deficit_hidrico_mm",
        y="Municipio",
        ax=ax,
        palette="Reds_r",
    )
    ax.set_title(
        "Municipios con Mayor Déficit Hídrico Climatológico Promedio (ET0 - Pef)",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Déficit Hídrico Promedio del Ciclo (mm)")
    ax.set_ylabel("Municipio")
    plt.tight_layout()

    filepath = OUTPUT_DIR / "deficit_hidrico_municipios.png"
    plt.savefig(filepath)
    plt.close()

    # Generar Mapa del Déficit Hídrico Climatológico
    mun_cve_map = df_indices[["Municipio", "CVE_MUN"]].drop_duplicates()
    df_deficit_mapped = df_deficit.merge(mun_cve_map, on="Municipio", how="left")
    merged_deficit = gdf.merge(df_deficit_mapped, on="CVE_MUN", how="inner")

    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    merged_deficit.plot(
        column="deficit_hidrico_mm",
        cmap="Reds",
        legend=True,
        legend_kwds={
            "label": "Déficit Hídrico Promedio (mm)",
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
    )
    ax.set_title(
        "Déficit Hídrico Climatológico Promedio (ET0 - Pef) por Municipio - Sonora",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    filepath_map = OUTPUT_DIR / "mapa_deficit_hidrico.png"
    plt.savefig(filepath_map, bbox_inches="tight")
    plt.close()

    # Calcular promedio estatal de fracción azul (dependencia de riego)
    mean_frac_azul = float(df_maestro["frac_azul"].mean() * 100)

    # 3. Guardar métricas clave en un archivo JSON estructurado para el sitio web
    summary_metrics = {
        "frecuencia_sequia_estatal_pct": round(pct_frec, 2),
        "frac_azul_estatal_pct": round(mean_frac_azul, 2),
        "municipio_mayor_deficit": df_deficit.iloc[0]["Municipio"],
        "max_deficit_mm": round(df_deficit.iloc[0]["deficit_hidrico_mm"], 2),
    }

    with open(
        BASE_DIR / "data" / "processed" / "resumen_sequia.json", "w", encoding="utf-8"
    ) as f:
        json.dump(summary_metrics, f, ensure_ascii=False, indent=2)

    # === ANÁLISIS EXTRA 1: ÍNDICE DE ARIDEZ (Ptot / ETo) ===
    df_aridez = (
        df_maestro.groupby("Municipio")
        .agg(
            {"Ptot_total_mm": "mean", "ET0_total_mm": "mean", "ratio_Ptot_ET0": "mean"}
        )
        .reset_index()
        .sort_values("ratio_Ptot_ET0", ascending=True)
    )

    # Graficar barra de aridez (municipios más áridos)
    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    sns.barplot(
        data=df_aridez.head(15),
        x="ratio_Ptot_ET0",
        y="Municipio",
        ax=ax,
        palette="YlOrBr",
    )
    ax.set_title(
        "Municipios con Mayor Grado de Aridez (Menor Ratio Ptot/ET0)",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Ratio Lluvia Total / Evapotranspiración (Ptot/ET0)")
    ax.set_ylabel("Municipio")
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "aridez_municipal.png")
    plt.close()

    # Generar Mapa del Índice de Aridez
    df_aridez_mapped = df_aridez.merge(mun_cve_map, on="Municipio", how="left")
    merged_aridez = gdf.merge(df_aridez_mapped, on="CVE_MUN", how="inner")
    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    merged_aridez.plot(
        column="ratio_Ptot_ET0",
        cmap="YlOrBr_r",
        legend=True,
        legend_kwds={
            "label": "Índice de Aridez (Ptot / ETo)",
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
    )
    ax.set_title(
        "Índice de Aridez Estacional Promedio (Ptot/ETo) por Municipio",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    plt.savefig(OUTPUT_DIR / "mapa_aridez_municipal.png", bbox_inches="tight")
    plt.close()

    # === ANÁLISIS EXTRA 2: EFICIENCIA DE LA LLUVIA (Pef vs Ptot) ===
    df_eficiencia = (
        df_maestro.groupby("Municipio")
        .agg({"Ptot_total_mm": "mean", "Pef_total_mm": "mean"})
        .reset_index()
        .sort_values("Ptot_total_mm", ascending=False)
    )

    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    df_ef_melted = df_eficiencia.head(15).melt(
        id_vars="Municipio",
        value_vars=["Ptot_total_mm", "Pef_total_mm"],
        var_name="Tipo",
        value_name="Precipitacion",
    )
    df_ef_melted["Tipo"] = df_ef_melted["Tipo"].map(
        {"Ptot_total_mm": "Lluvia Total", "Pef_total_mm": "Lluvia Efectiva"}
    )
    sns.barplot(
        data=df_ef_melted,
        x="Precipitacion",
        y="Municipio",
        hue="Tipo",
        ax=ax,
        palette="Blues_r",
    )
    ax.set_title(
        "Lluvia Total vs. Lluvia Efectiva por Municipio", pad=15, weight="bold"
    )
    ax.set_xlabel("Precipitación Acumulada del Ciclo (mm)")
    ax.set_ylabel("Municipio")
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eficiencia_lluvia_municipios.png")
    plt.close()

    # === ANÁLISIS EXTRA 3: DÍAS DE LLUVIA SIGNIFICATIVA ===
    df_dias = (
        df_maestro.groupby("Municipio")
        .agg({"dias_lluvia_ge_1mm": "mean", "dias_lluvia_ge_5mm": "mean"})
        .reset_index()
        .sort_values("dias_lluvia_ge_1mm", ascending=False)
    )

    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    df_dias_melted = df_dias.head(15).melt(
        id_vars="Municipio",
        value_vars=["dias_lluvia_ge_1mm", "dias_lluvia_ge_5mm"],
        var_name="Umbral",
        value_name="Dias",
    )
    df_dias_melted["Umbral"] = df_dias_melted["Umbral"].map(
        {"dias_lluvia_ge_1mm": "Lluvia ≥ 1mm", "dias_lluvia_ge_5mm": "Lluvia ≥ 5mm"}
    )
    sns.barplot(
        data=df_dias_melted,
        x="Dias",
        y="Municipio",
        hue="Umbral",
        ax=ax,
        palette="Purples_r",
    )
    ax.set_title(
        "Días Promedio de Lluvia Significativa en el Ciclo por Municipio",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Número de Días")
    ax.set_ylabel("Municipio")
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "dias_lluvia_municipios.png")
    plt.close()

    # === SERIE HISTÓRICA 1: EVOLUCIÓN TEMPORAL DE SEQUÍA (2003-2026) ===
    df_pct_grouped = df_mon.groupby(["Fecha", "Categoria"]).size().unstack(fill_value=0)
    df_pct = (df_pct_grouped / 72 * 100).clip(upper=100)
    df_pct["Sin Sequía"] = (100 - df_pct.sum(axis=1)).clip(lower=0)
    cols_order = ["Sin Sequía", "D0", "D1", "D2", "D3", "D4"]
    df_pct = df_pct[cols_order]

    colors_drought = ["#e2e8f0", "#ffd966", "#f6b26b", "#e06666", "#cc0000", "#7f0000"]
    labels_drought = [
        "Sin Sequía",
        "D0: Anormalmente Seco",
        "D1: Sequía Moderada",
        "D2: Sequía Severa",
        "D3: Sequía Extrema",
        "D4: Sequía Excepcional",
    ]

    fig, ax = plt.subplots(figsize=(12, 6), dpi=300)
    ax.stackplot(
        df_pct.index, df_pct.T, labels=labels_drought, colors=colors_drought, alpha=0.85
    )
    ax.set_title(
        "Evolución Histórica de la Severidad de la Sequía en Sonora (2003-2026)",
        pad=15,
        weight="bold",
    )
    ax.set_ylabel("Porcentaje de Municipios (%)")
    ax.set_xlabel("Año")
    ax.set_xlim(df_pct.index.min(), df_pct.index.max())
    ax.set_ylim(0, 100)
    ax.legend(
        loc="upper left",
        bbox_to_anchor=(1.01, 1),
        title="Categorías de Sequía",
        frameon=True,
        shadow=True,
    )
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "evolucion_sequia_historica.png", bbox_inches="tight")
    plt.close()

    # === SERIE HISTÓRICA 2: TENDENCIA TEMPERATURA VS LLUVIA (2010-2024) ===
    df_yearly = (
        df_maestro.groupby("Anio")
        .agg({"Ptot_total_mm": "mean", "Tmean_ciclo_mean_C": "mean"})
        .reset_index()
    )

    fig, ax1 = plt.subplots(figsize=(10, 6), dpi=300)
    color_blue = "tab:blue"
    ax1.set_xlabel("Año")
    ax1.set_ylabel("Precipitación Total Promedio (mm)", color=color_blue)
    ax1.bar(
        df_yearly["Anio"],
        df_yearly["Ptot_total_mm"],
        color=color_blue,
        alpha=0.5,
        label="Precipitación",
    )
    ax1.tick_params(axis="y", labelcolor=color_blue)

    ax2 = ax1.twinx()
    color_red = "tab:red"
    ax2.set_ylabel("Temperatura Media Promedio (°C)", color=color_red)
    ax2.plot(
        df_yearly["Anio"],
        df_yearly["Tmean_ciclo_mean_C"],
        color=color_red,
        marker="o",
        linewidth=2.5,
        label="Temperatura",
    )
    ax2.tick_params(axis="y", labelcolor=color_red)

    ax1.set_title(
        "Tendencia Histórica de Precipitación y Temperatura Promedio en Sonora (2010-2023)",
        pad=15,
        weight="bold",
    )
    ax1.set_xticks(df_yearly["Anio"])
    fig.tight_layout()
    plt.savefig(OUTPUT_DIR / "tendencia_climatica_anual.png")
    plt.close()

    # === ANÁLISIS EXTRA 4: ÍNDICE DE DEPENDENCIA DE RIEGO (FRACCIÓN AZUL VS VERDE) ===
    df_frac = (
        df_maestro.groupby("Municipio")
        .agg({"frac_azul": "mean", "frac_verde": "mean"})
        .reset_index()
        .sort_values("frac_azul", ascending=False)
    )

    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    ax.barh(
        df_frac.head(15)["Municipio"],
        df_frac.head(15)["frac_azul"] * 100,
        label="Fracción Azul (Riego)",
        color="#38bdf8",
        alpha=0.9,
    )
    ax.barh(
        df_frac.head(15)["Municipio"],
        df_frac.head(15)["frac_verde"] * 100,
        left=df_frac.head(15)["frac_azul"] * 100,
        label="Fracción Verde (Lluvia)",
        color="#a3e635",
        alpha=0.9,
    )
    ax.set_title(
        "Dependencia de Riego vs. Lluvia Efectiva (Fracción Azul vs. Verde)",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Porcentaje (%)")
    ax.set_ylabel("Municipio")
    ax.legend(loc="lower left")
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "fraccion_azul_verde_municipios.png")
    plt.close()

    # === ANÁLISIS EXTRA 5: ESTRÉS TÉRMICO Y RIESGO AGROCLIMÁTICO ===
    df_estres = (
        df_maestro.groupby("Municipio")
        .agg({"dias_estres_calor": "mean", "dias_riesgo_helada": "mean"})
        .reset_index()
        .sort_values("dias_estres_calor", ascending=False)
    )

    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    df_estres_melted = df_estres.head(15).melt(
        id_vars="Municipio",
        value_vars=["dias_estres_calor", "dias_riesgo_helada"],
        var_name="Riesgo",
        value_name="Dias",
    )
    df_estres_melted["Riesgo"] = df_estres_melted["Riesgo"].map(
        {
            "dias_estres_calor": "Días Estrés Calor",
            "dias_riesgo_helada": "Días Riesgo Helada",
        }
    )
    sns.barplot(
        data=df_estres_melted,
        x="Dias",
        y="Municipio",
        hue="Riesgo",
        ax=ax,
        palette="coolwarm",
    )
    ax.set_title(
        "Estrés Agroclimático por Temperaturas Extremas (Días por Ciclo)",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Número Promedio de Días")
    ax.set_ylabel("Municipio")
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "estres_termico_municipios.png")
    plt.close()

    # === 3 MAPAS DE TEMPERATURA POR MUNICIPIO ===
    df_temp = (
        df_maestro.groupby("Municipio")
        .agg(
            {
                "Tmean_ciclo_mean_C": "mean",
                "Tmax_ciclo_mean_C": "mean",
                "Tmin_ciclo_mean_C": "mean",
            }
        )
        .reset_index()
    )

    df_temp_mapped = df_temp.merge(mun_cve_map, on="Municipio", how="left")
    merged_temp = gdf.merge(df_temp_mapped, on="CVE_MUN", how="inner")

    # 1. Mapa Temperatura Media
    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    merged_temp.plot(
        column="Tmean_ciclo_mean_C",
        cmap="Oranges",
        legend=True,
        legend_kwds={
            "label": "Temperatura Media (°C)",
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
    )
    ax.set_title(
        "Temperatura Media Promedio del Ciclo por Municipio - Sonora",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    plt.savefig(OUTPUT_DIR / "mapa_temperatura_media.png", bbox_inches="tight")
    plt.close()

    # 2. Mapa Temperatura Máxima
    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    merged_temp.plot(
        column="Tmax_ciclo_mean_C",
        cmap="Reds",
        legend=True,
        legend_kwds={
            "label": "Temperatura Máxima (°C)",
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
    )
    ax.set_title(
        "Temperatura Máxima Promedio del Ciclo por Municipio - Sonora",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    plt.savefig(OUTPUT_DIR / "mapa_temperatura_maxima.png", bbox_inches="tight")
    plt.close()

    # 3. Mapa Temperatura Mínima
    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    merged_temp.plot(
        column="Tmin_ciclo_mean_C",
        cmap="Blues",
        legend=True,
        legend_kwds={
            "label": "Temperatura Mínima (°C)",
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
    )
    ax.set_title(
        "Temperatura Mínima Promedio del Ciclo por Municipio - Sonora",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    plt.savefig(OUTPUT_DIR / "mapa_temperatura_minima.png", bbox_inches="tight")
    plt.close()

    print(
        "Métricas agregadas y gráficas de déficit hídrico, precipitación y temperatura generadas exitosamente."
    )


def analyze_dam_volumes():
    """Extrae y grafica la evolución histórica del volumen bruto
    de agua utilizado por los Distritos de Riego de Sonora.
    Fuente: CONAGUA, Superficies Regadas y Volúmenes 1998-2024.
    """
    import openpyxl

    xlsx_path = (
        BASE_DIR
        / "data"
        / "raw"
        / "conagua"
        / "SUPERFICIES_REGADAS_Y_VOL_MENES_1998-2024.xlsx"
    )
    if not xlsx_path.exists():
        print(f"Archivo CONAGUA no encontrado: {xlsx_path}")
        return

    sonora_dists = {
        "018": "DR-018 Col. Yaquis",
        "037": "DR-037 Altar-Pitiquito",
        "038": "DR-038 Río Mayo",
        "041": "DR-041 Río Yaqui",
        "051": "DR-051 Costa Hermosillo",
    }

    records = []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    for year_str in wb.sheetnames:
        year = int(year_str)
        ws = wb[year_str]
        rows = list(ws.iter_rows(min_row=1, max_row=600, values_only=True))
        for i, row in enumerate(rows):
            code = str(row[0]).strip() if row[0] else ""
            if code in sonora_dists:
                if i + 2 < len(rows):
                    vol_row = rows[i + 2]
                    total_vol = vol_row[16] if len(vol_row) > 16 else vol_row[-1]
                    if total_vol:
                        try:
                            vol = float(
                                str(total_vol).replace(",", "").replace(" ", "")
                            )
                            records.append(
                                {
                                    "Año": year,
                                    "DR": sonora_dists[code],
                                    "Volumen_Mm3": round(vol / 1000, 2),
                                }
                            )
                        except (ValueError, TypeError):
                            pass
    wb.close()

    df = pd.DataFrame(records)
    if df.empty:
        print("No se encontraron datos de volumen de DR.")
        return

    # --- Gráfica 1: Líneas por DR ---
    fig, ax = plt.subplots(figsize=(12, 6), dpi=300)
    palette = ["#2563eb", "#dc2626", "#16a34a", "#f59e0b", "#8b5cf6"]
    for idx, (dr_name, grp) in enumerate(df.sort_values("Año").groupby("DR")):
        ax.plot(
            grp["Año"],
            grp["Volumen_Mm3"],
            marker="o",
            markersize=4,
            linewidth=2,
            label=dr_name,
            color=palette[idx % len(palette)],
        )
    ax.set_title(
        "Evolución del Volumen Bruto de Riego por Distrito "
        "de Riego en Sonora (1998-2024)",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Año")
    ax.set_ylabel("Volumen Bruto (Millones de m³)")
    ax.legend(
        loc="upper left",
        bbox_to_anchor=(1.01, 1),
        title="Distritos de Riego",
        frameon=True,
        shadow=True,
    )
    ax.set_xlim(1997, 2025)
    plt.tight_layout()
    plt.savefig(
        OUTPUT_DIR / "evolucion_volumen_dr.png",
        bbox_inches="tight",
    )
    plt.close()

    # --- Gráfica 2: Área apilada del volumen total ---
    pivot = df.pivot_table(index="Año", columns="DR", values="Volumen_Mm3").fillna(0)
    pivot = pivot.sort_index()

    fig, ax = plt.subplots(figsize=(12, 6), dpi=300)
    ax.stackplot(
        pivot.index,
        pivot.T,
        labels=pivot.columns,
        colors=palette[: len(pivot.columns)],
        alpha=0.8,
    )
    ax.set_title(
        "Volumen Total de Agua para Riego en Sonora por Distrito de Riego (1998-2024)",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Año")
    ax.set_ylabel("Volumen Bruto Acumulado (Mm³)")
    ax.legend(
        loc="upper left",
        bbox_to_anchor=(1.01, 1),
        title="Distritos de Riego",
        frameon=True,
        shadow=True,
    )
    ax.set_xlim(1997, 2025)
    plt.tight_layout()
    plt.savefig(
        OUTPUT_DIR / "volumen_total_apilado_dr.png",
        bbox_inches="tight",
    )
    plt.close()

    print("Gráficas de evolución de volumen de riego por DR generadas exitosamente.")


def compute_composite_vulnerability(gdf, df_indices):
    """Calcula un índice compuesto de vulnerabilidad hídrica
    municipal integrando múltiples indicadores normalizados.
    Genera un mapa coroplético de vulnerabilidad integral.
    """
    df_maestro = pd.read_csv(MAESTRO_CSV)

    # Agregar métricas por municipio
    df_mun = (
        df_maestro.groupby("Municipio")
        .agg(
            {
                "deficit_hidrico_mm": "mean",
                "ratio_Ptot_ET0": "mean",
                "frac_azul": "mean",
                "dias_estres_calor": "mean",
                "dias_riesgo_helada": "mean",
            }
        )
        .reset_index()
    )

    # Merge con ISAG (invertir: ISAG 1=Alta vuln, 3=Baja)
    df_isag = df_indices[["Municipio", "CVE_MUN", "ISAG_PV", "ISAG_OI"]].copy()
    df_isag["ISAG_mean"] = (df_isag["ISAG_PV"] + df_isag["ISAG_OI"]) / 2
    # Invertir: menor ISAG = mayor vulnerabilidad
    df_isag["vuln_isag"] = 1 - (
        (df_isag["ISAG_mean"] - 1) / 2
    )  # 1->1.0, 2->0.5, 3->0.0

    df_merged = df_mun.merge(
        df_isag[["Municipio", "CVE_MUN", "vuln_isag"]],
        on="Municipio",
        how="inner",
    )

    # Normalización min-max [0, 1] de cada indicador
    def norm_minmax(series):
        mn, mx = series.min(), series.max()
        if mx - mn == 0:
            return series * 0
        return (series - mn) / (mx - mn)

    df_merged["n_deficit"] = norm_minmax(df_merged["deficit_hidrico_mm"])
    # Invertir aridez: menor ratio = mayor vuln
    df_merged["n_aridez"] = 1 - norm_minmax(df_merged["ratio_Ptot_ET0"])
    df_merged["n_frac_azul"] = norm_minmax(df_merged["frac_azul"])
    df_merged["n_estres"] = norm_minmax(
        df_merged["dias_estres_calor"] + df_merged["dias_riesgo_helada"]
    )
    df_merged["n_isag"] = norm_minmax(df_merged["vuln_isag"])

    # Índice compuesto ponderado
    weights = {
        "n_deficit": 0.35,
        "n_isag": 0.20,
        "n_frac_azul": 0.35,
        "n_estres": 0.10,
    }
    df_merged["IVH"] = sum(df_merged[col] * w for col, w in weights.items())

    # Clasificar en categorías
    def classify_ivh(val):
        if val >= 0.70:
            return "Crítica"
        elif val >= 0.50:
            return "Alta"
        elif val >= 0.30:
            return "Moderada"
        else:
            return "Baja"

    df_merged["IVH_cat"] = df_merged["IVH"].apply(classify_ivh)

    # --- MAPA COROPLÉTICO CONTINUO ---
    geo_merged = gdf.merge(df_merged, on="CVE_MUN", how="inner")

    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    geo_merged.plot(
        column="IVH",
        cmap="RdYlGn_r",
        legend=True,
        legend_kwds={
            "label": ("Índice de Vulnerabilidad Hídrica (IVH)"),
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
        vmin=0,
        vmax=1,
    )
    ax.set_title(
        "Índice Compuesto de Vulnerabilidad Hídrica Municipal — Sonora",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    plt.savefig(
        OUTPUT_DIR / "mapa_vulnerabilidad_compuesta.png",
        bbox_inches="tight",
    )
    plt.close()

    # --- GRÁFICA DE BARRAS TOP 15 ---
    df_top = df_merged.sort_values("IVH", ascending=False).head(15)

    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    colors_bar = [
        "#dc2626"
        if v >= 0.70
        else "#f59e0b"
        if v >= 0.50
        else "#22c55e"
        if v >= 0.30
        else "#3b82f6"
        for v in df_top["IVH"]
    ]
    ax.barh(
        df_top["Municipio"],
        df_top["IVH"],
        color=colors_bar,
        edgecolor="#1e293b",
        linewidth=0.5,
    )
    ax.set_title(
        "Municipios con Mayor Vulnerabilidad Hídrica Integral (IVH)",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Índice de Vulnerabilidad Hídrica (0-1)")
    ax.set_ylabel("Municipio")
    ax.set_xlim(0, 1)

    # Leyenda manual
    from matplotlib.patches import Patch

    legend_elements = [
        Patch(
            facecolor="#dc2626",
            edgecolor="black",
            label="Crítica (≥0.70)",
        ),
        Patch(
            facecolor="#f59e0b",
            edgecolor="black",
            label="Alta (0.50-0.69)",
        ),
        Patch(
            facecolor="#22c55e",
            edgecolor="black",
            label="Moderada (0.30-0.49)",
        ),
        Patch(
            facecolor="#3b82f6",
            edgecolor="black",
            label="Baja (<0.30)",
        ),
    ]
    ax.legend(
        handles=legend_elements,
        title="Categoría IVH",
        loc="lower right",
        frameon=True,
        shadow=True,
    )
    plt.tight_layout()
    plt.savefig(
        OUTPUT_DIR / "vulnerabilidad_ranking_municipios.png",
        bbox_inches="tight",
    )
    plt.close()

    # --- GRÁFICA RADAR/DESGLOSE TOP 5 ---
    df_radar = df_merged.sort_values("IVH", ascending=False).head(5)
    components = [
        "n_deficit",
        "n_isag",
        "n_frac_azul",
        "n_estres",
    ]
    comp_labels = [
        "Déficit\nHídrico",
        "Sequía\n(ISAG)",
        "Dependencia\nRiego",
        "Estrés\nTérmico",
    ]

    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    x_pos = range(len(components))
    bar_width = 0.15
    palette_radar = [
        "#ef4444",
        "#f97316",
        "#eab308",
        "#22c55e",
        "#3b82f6",
    ]

    for i, (_, row) in enumerate(df_radar.iterrows()):
        offset = (i - 2) * bar_width
        vals = [row[c] for c in components]
        ax.bar(
            [p + offset for p in x_pos],
            vals,
            bar_width,
            label=row["Municipio"],
            color=palette_radar[i],
            alpha=0.85,
            edgecolor="#1e293b",
            linewidth=0.3,
        )

    ax.set_title(
        "Desglose de Vulnerabilidad por Componente — Top 5 Municipios",
        pad=15,
        weight="bold",
    )
    ax.set_xlabel("Componente del Índice")
    ax.set_ylabel("Valor Normalizado (0-1)")
    ax.set_xticks(x_pos)
    ax.set_xticklabels(comp_labels)
    ax.legend(
        title="Municipio",
        loc="upper right",
        frameon=True,
        shadow=True,
    )
    ax.set_ylim(0, 1.1)
    plt.tight_layout()
    plt.savefig(
        OUTPUT_DIR / "vulnerabilidad_desglose_top5.png",
        bbox_inches="tight",
    )
    plt.close()

    print("Mapa y gráficas de vulnerabilidad hídrica compuesta generados exitosamente.")


def generate_ddr_technification_map(gdf, df_indices):
    """Genera mapa coroplético de la tecnificación del riego por DDR."""
    siap_2024_path = (
        BASE_DIR
        / "data"
        / "processed"
        / "siap_produccion"
        / "sonora"
        / "cierre_agricola_sonora_2024.csv"
    )
    if not siap_2024_path.exists():
        print("No se encontró cierre_agricola_sonora_2024.csv, abortando.")
        return

    siap = pd.read_csv(siap_2024_path)
    mun_ddr_map = siap[["Nommunicipio", "Idddr"]].drop_duplicates()
    mun_ddr_map["Nommunicipio"] = mun_ddr_map["Nommunicipio"].str.strip()

    with open(CODIF_JSON, encoding="utf-8") as f:
        codif = json.load(f)
    ddr_names = codif["codigos_ddrs"]
    mun_ddr_map["DDR"] = mun_ddr_map["Idddr"].astype(str).map(ddr_names)

    # Limpiar columna Municipio en gdf para asegurar coincidencia
    gdf_temp = gdf.copy()
    gdf_temp["Municipio_clean"] = gdf_temp["Municipio"].str.strip()

    # Unir con el mapa de DDR
    gdf_merged = gdf_temp.merge(
        mun_ddr_map, left_on="Municipio_clean", right_on="Nommunicipio", how="left"
    )

    # Diccionario de porcentaje de tecnificación (SAGARPA/UNISON 2022)
    tecnif_map = {
        "Guaymas": 78.0,
        "Caborca": 70.0,
        "Hermosillo": 47.0,
        "Mazatán": 16.0,
        "Magdalena": 14.0,
        "Navojoa": 13.0,
        "San Luis Río Colorado": 12.0,
        "Ures": 12.0,
        "Cajeme": 8.0,
        "Moctezuma": 7.0,
        "Agua Prieta": 3.0,
        "Sahuaripa": 2.0,
    }

    gdf_merged["Tecnificacion"] = gdf_merged["DDR"].map(tecnif_map).fillna(0)

    fig, ax = plt.subplots(figsize=(10, 8), dpi=300)

    # Dibujar el fondo del mapa (municipios)
    gdf_merged.plot(
        column="Tecnificacion",
        cmap="YlGn",
        edgecolor="#1e293b",
        linewidth=0.4,
        legend=True,
        legend_kwds={
            "label": "Porcentaje de Tecnificación (%)",
            "orientation": "horizontal",
            "shrink": 0.7,
            "pad": 0.05,
        },
        ax=ax,
        missing_kwds={"color": "lightgrey", "edgecolor": "black", "label": "Sin Datos"},
    )

    # Anotar los nombres de los DDR en sus centroides
    gdf_ddr = gdf_merged.dropna(subset=["DDR"]).dissolve(by="DDR")
    for idx, row in gdf_ddr.iterrows():
        if pd.isna(idx) or idx not in tecnif_map:
            continue
        centroid = row.geometry.centroid
        ax.annotate(
            text=f"{idx}\n({tecnif_map[idx]:.0f}%)",
            xy=(centroid.x, centroid.y),
            xytext=(0, 0),
            textcoords="offset points",
            fontsize=8,
            weight="bold",
            color="#0f172a",
            ha="center",
            va="center",
            bbox={
                "boxstyle": "round,pad=0.2",
                "fc": "white",
                "alpha": 0.7,
                "ec": "none",
            },
        )

    ax.set_title(
        "Distribución Espacial de la Tecnificación del Riego por DDR\n",
        pad=15,
        weight="bold",
    )
def generate_efficiency_productivity_maps(gdf, df_indices):
    """Genera mapas coropléticos municipales de eficiencia física y productividad económica."""
    df_maestro = pd.read_csv(MAESTRO_CSV)
    df_maestro = df_maestro[df_maestro["SupCosechadaTotal_ha"] > 0].copy()
    
    # Calcular volumen de agua total
    df_maestro["vol_agua"] = df_maestro["UACtotal_m3_ha"] * df_maestro["SupCosechadaTotal_ha"]
    
    df_mun = df_maestro.groupby("Municipio").agg({
        "VolumenTotal_t": "sum",
        "vol_agua": "sum",
        "Valorproduccion": "sum"
    }).reset_index()
    
    # Filtrar municipios con volumen de agua > 0
    df_mun = df_mun[df_mun["vol_agua"] > 0].copy()
    
    # Eficiencia física (Ton/m3)
    df_mun["eficiencia_fisica"] = df_mun["VolumenTotal_t"] / df_mun["vol_agua"]
    # Productividad económica (MXN/m3)
    df_mun["productividad_economica"] = df_mun["Valorproduccion"] / df_mun["vol_agua"]
    
    # Cruzar con CVE_MUN
    mun_cve_map = df_indices[["Municipio", "CVE_MUN"]].drop_duplicates().copy()
    df_mun["Municipio_clean"] = df_mun["Municipio"].str.strip()
    mun_cve_map["Municipio_clean"] = mun_cve_map["Municipio"].str.strip()
    
    df_mun_mapped = df_mun.merge(mun_cve_map[["Municipio_clean", "CVE_MUN"]], on="Municipio_clean", how="left")
    merged_gdf = gdf.merge(df_mun_mapped, on="CVE_MUN", how="inner")
    
    # 1. Mapa de Eficiencia Física
    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    merged_gdf.plot(
        column="eficiencia_fisica",
        cmap="YlGnBu",
        legend=True,
        legend_kwds={
            "label": "Eficiencia Hídrica Física (Ton/m³)",
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
    )
    ax.set_title(
        "Eficiencia Hídrica Física Promedio por Municipio - Sonora",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    plt.savefig(OUTPUT_DIR / "mapa_eficiencia_fisica_municipal.png", bbox_inches="tight")
    plt.close()
    
    # 2. Mapa de Productividad Económica
    fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=300)
    merged_gdf.plot(
        column="productividad_economica",
        cmap="YlOrRd",
        legend=True,
        legend_kwds={
            "label": "Productividad Económica (MXN/m³)",
            "orientation": "horizontal",
            "pad": 0.05,
            "shrink": 0.7,
        },
        edgecolor="black",
        linewidth=0.4,
        ax=ax,
    )
    ax.set_title(
        "Productividad Económica del Agua Promedio por Municipio - Sonora",
        pad=15,
        weight="bold",
    )
    ax.axis("off")
    plt.savefig(OUTPUT_DIR / "mapa_productividad_economica_municipal.png", bbox_inches="tight")
    plt.close()
    
    print("Mapas de eficiencia hídrica física y productividad económica generados exitosamente.")


def plot_ranking_hh():
    """Genera gráfica de ranking de HH por cultivo (top eficientes y bottom ineficientes)."""
    df_maestro = pd.read_csv(MAESTRO_CSV)
    crop_hh = (
        df_maestro.groupby("Cultivo")
        .agg(
            {
                "HHtotal_m3_ton": "mean",
                "HHazul_m3_ton": "mean",
                "HHverde_m3_ton": "mean",
                "Rend_t_ha": "mean",
                "PMR": "mean",
                "VolumenTotal_t": "sum",
            }
        )
        .reset_index()
        .sort_values("HHtotal_m3_ton", ascending=True)
    )
    crop_hh = crop_hh[crop_hh["VolumenTotal_t"] > 0]

    top_n = 20
    bot_n = 15

    top_crops = crop_hh.head(top_n).copy()
    bot_crops = crop_hh.tail(bot_n).copy()
    bot_crops = bot_crops.sort_values("HHtotal_m3_ton", ascending=False)

    for label, data, filename in [
        ("Top 20 Cultivos con Menor Huella Hídrica", top_crops, "ranking_hh_top20.png"),
        ("Top 15 Cultivos con Mayor Huella Hídrica", bot_crops, "ranking_hh_bottom15.png"),
    ]:
        fig, ax = plt.subplots(figsize=(12, 7), dpi=300)
        y_pos = range(len(data))
        ax.barh(
            y_pos,
            data["HHazul_m3_ton"].values,
            color="#38bdf8",
            label="HH Azul (riego)",
            alpha=0.9,
        )
        ax.barh(
            y_pos,
            data["HHverde_m3_ton"].values,
            left=data["HHazul_m3_ton"].values,
            color="#4ade80",
            label="HH Verde (lluvia)",
            alpha=0.9,
        )
        ax.set_yticks(y_pos)
        ax.set_yticklabels(data["Cultivo"].values, fontsize=9)
        ax.set_xlabel("Huella Hídrica Total (m³/ton)")
        ax.set_title(label, pad=15, weight="bold")
        ax.legend(loc="lower right", frameon=True, shadow=True)
        ax.invert_yaxis()
        plt.tight_layout()
        plt.savefig(OUTPUT_DIR / filename, bbox_inches="tight")
        plt.close()

    print("Ranking de HH por cultivo generado exitosamente.")


def compute_ddr_summary_table():
    """Genera tabla resumen consolidada por DDR con métricas clave."""
    df_maestro = pd.read_csv(MAESTRO_CSV)
    df_maestro = df_maestro[df_maestro["SupCosechadaTotal_ha"] > 0].copy()
    df_maestro["vol_agua"] = df_maestro["UACtotal_m3_ha"] * df_maestro["SupCosechadaTotal_ha"]

    ddr_agg = (
        df_maestro.groupby("DDR")
        .agg(
            {
                "HHtotal_m3_ton": "mean",
                "HHazul_m3_ton": "mean",
                "frac_azul": "mean",
                "deficit_hidrico_mm": "mean",
                "Rend_t_ha": "mean",
                "VolumenTotal_t": "sum",
                "Valorproduccion": "sum",
                "vol_agua": "sum",
                "SupCosechadaTotal_ha": "sum",
                "PMR": "mean",
            }
        )
        .reset_index()
    )
    ddr_agg["prod_economica_mxn_m3"] = ddr_agg["Valorproduccion"] / ddr_agg["vol_agua"]
    ddr_agg["RAV_L_mxn"] = ddr_agg["vol_agua"] * 1000 / ddr_agg["Valorproduccion"]

    tecnif_map = {
        "Guaymas": 78.0, "Caborca": 70.0, "Hermosillo": 47.0,
        "Mazatán": 16.0, "Magdalena": 14.0, "Navojoa": 13.0,
        "San Luis Río Colorado": 12.0, "Ures": 12.0,
        "Cajeme": 8.0, "Moctezuma": 7.0, "Agua Prieta": 3.0, "Sahuaripa": 2.0,
    }
    ddr_agg["tecnificacion_pct"] = ddr_agg["DDR"].map(tecnif_map)

    ddr_agg = ddr_agg.sort_values("RAV_L_mxn", ascending=True)

    output_csv = OUTPUT_DIR / "ddr_summary.csv"
    ddr_agg.to_csv(output_csv, index=False, encoding="utf-8")
    print(f"Tabla resumen DDR guardada en: {output_csv}")
    return ddr_agg


def plot_cuadrante_hh_pmr():
    """Genera gráfica de cuadrante HH vs PMR para todos los cultivos."""
    df_maestro = pd.read_csv(MAESTRO_CSV)
    crop_agg = (
        df_maestro.groupby("Cultivo")
        .agg({"HHtotal_m3_ton": "mean", "PMR": "mean", "VolumenTotal_t": "sum"})
        .reset_index()
    )
    crop_agg = crop_agg[crop_agg["VolumenTotal_t"] > 0].copy()

    hh_median = crop_agg["HHtotal_m3_ton"].median()
    pmr_median = crop_agg["PMR"].median()

    fig, ax = plt.subplots(figsize=(14, 10), dpi=300)

    cuadrantes = {
        "ideal": (crop_agg["HHtotal_m3_ton"] <= hh_median) & (crop_agg["PMR"] >= pmr_median),
        "ineficiente": (crop_agg["HHtotal_m3_ton"] > hh_median) & (crop_agg["PMR"] < pmr_median),
        "alta_hh_alto_pmr": (crop_agg["HHtotal_m3_ton"] > hh_median) & (crop_agg["PMR"] >= pmr_median),
        "baja_hh_bajo_pmr": (crop_agg["HHtotal_m3_ton"] <= hh_median) & (crop_agg["PMR"] < pmr_median),
    }

    colors = {"ideal": "#22c55e", "ineficiente": "#ef4444", "alta_hh_alto_pmr": "#f59e0b", "baja_hh_bajo_pmr": "#3b82f6"}
    labels_map = {
        "ideal": "Baja HH + Alto PMR (Ideal)",
        "ineficiente": "Alta HH + Bajo PMR (Problemático)",
        "alta_hh_alto_pmr": "Alta HH + Alto PMR",
        "baja_hh_bajo_pmr": "Baja HH + Bajo PMR",
    }

    for key, mask in cuadrantes.items():
        subset = crop_agg[mask]
        ax.scatter(
            subset["HHtotal_m3_ton"], subset["PMR"],
            c=colors[key], label=labels_map[key], alpha=0.7, s=50, edgecolors="#333", linewidths=0.5,
        )

    annotate_crops = [
        "Trigo grano", "Maíz grano", "Pepino", "Tomate rojo (jitomate)", "Papa",
        "Alfalfa achicalada", "Sorgo grano", "Frijol", "Uva", "Nuez", "Cártamo",
        "Cebolla", "Espárrago", "Sandía", "Melón", "Naranja", "Agave",
    ]
    for crop in annotate_crops:
        row = crop_agg[crop_agg["Cultivo"] == crop]
        if not row.empty:
            xx = row["HHtotal_m3_ton"].values[0]
            yy = row["PMR"].values[0]
            ax.annotate(
                crop, (xx, yy), fontsize=7, fontweight="bold",
                xytext=(5, 5), textcoords="offset points",
                bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="gray", alpha=0.8),
            )

    ax.axvline(x=hh_median, color="gray", linestyle="--", alpha=0.5, label=f"Mediana HH = {hh_median:.0f}")
    ax.axhline(y=pmr_median, color="gray", linestyle=":", alpha=0.5, label=f"Mediana PMR = ${pmr_median:,.0f}")
    ax.set_xscale("log")
    ax.set_xlabel("Huella Hídrica Total (m³/ton) — escala log", fontsize=11)
    ax.set_ylabel("Precio Medio Rural ($/ton)", fontsize=11)
    ax.set_title("Cuadrante Eficiencia Hídrica vs. Rentabilidad por Cultivo", pad=15, weight="bold", fontsize=13)
    ax.legend(loc="upper left", frameon=True, shadow=True, fontsize=9)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "cuadrante_rentabilidad_eficiencia.png", bbox_inches="tight")
    plt.close()
    print("Cuadrante HH vs PMR generado exitosamente.")


def plot_ddr_cajeme_profile():
    """Genera figuras del perfil hídrico-productivo del DDR Cajeme."""
    df_maestro = pd.read_csv(MAESTRO_CSV)

    cajeme = df_maestro[df_maestro["Municipio"] == "Cajeme"].copy()
    cajeme = cajeme[cajeme["VolumenTotal_t"] > 0]

    # HH por cultivo en Cajeme (top 15)
    crops_cajeme = (
        cajeme.groupby("Cultivo")
        .agg({
            "HHtotal_m3_ton": "mean",
            "HHazul_m3_ton": "mean",
            "HHverde_m3_ton": "mean",
            "VolumenTotal_t": "sum",
            "PMR": "mean",
        })
        .sort_values("VolumenTotal_t", ascending=False)
        .head(15)
        .sort_values("HHtotal_m3_ton", ascending=True)
    )

    fig, ax = plt.subplots(figsize=(12, 7), dpi=300)
    y_pos = range(len(crops_cajeme))
    ax.barh(y_pos, crops_cajeme["HHazul_m3_ton"].values, color="#38bdf8", label="HH Azul", alpha=0.9)
    ax.barh(
        y_pos, crops_cajeme["HHverde_m3_ton"].values,
        left=crops_cajeme["HHazul_m3_ton"].values, color="#4ade80", label="HH Verde", alpha=0.9,
    )
    ax.set_yticks(y_pos)
    ax.set_yticklabels(crops_cajeme.index, fontsize=9)
    ax.set_xlabel("Huella Hídrica (m³/ton)")
    ax.set_title("Principales Cultivos en Cajeme: Huella Hídrica por Tonelada", pad=15, weight="bold")
    ax.legend(loc="lower right", frameon=True, shadow=True)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "cajeme_hh_cultivos.png", bbox_inches="tight")
    plt.close()

    # RAV por cultivo en Cajeme: top 10 best and worst
    with open(REPORTS_DIR / "fase3" / "prueba_analisis" / "rav_results.json", encoding="utf-8") as f:
        rav_data = json.load(f)
    cajeme_rav = rav_data["ddr_crops"]["Cajeme"]
    best = cajeme_rav["best"]
    worst = cajeme_rav["worst"]

    # Combine for a single horizontal bar chart
    labels_rev = [d["Cultivo"] for d in reversed(best)] + [d["Cultivo"] for d in worst]
    values_rev = [d["RAV_L_mxn"] for d in reversed(best)] + [d["RAV_L_mxn"] for d in worst]
    colors_rev = ["#22c55e"] * len(best) + ["#ef4444"] * len(worst)

    fig, ax = plt.subplots(figsize=(12, 6), dpi=300)
    y_pos = range(len(labels_rev))
    ax.barh(y_pos, values_rev, color=colors_rev, alpha=0.9, edgecolor="#333", linewidth=0.5)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels_rev, fontsize=9)
    ax.set_xlabel("Relación Agua-Valor (L/MXN)")
    ax.set_title("RAV por Cultivo en Cajeme: Más Eficientes (verde) vs. Menos Eficientes (rojo)", pad=15, weight="bold")

    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor="#22c55e", edgecolor="black", label="Más Eficientes (Menor RAV)"),
        Patch(facecolor="#ef4444", edgecolor="black", label="Menos Eficientes (Mayor RAV)"),
    ]
    ax.legend(handles=legend_elements, loc="lower right", frameon=True, shadow=True)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "cajeme_rav_cultivos.png", bbox_inches="tight")
    plt.close()

    print("Perfil de Cajeme generado exitosamente.")


def plot_trigo_vs_alternatives():
    """Genera gráfica comparativa de trigo grano vs. cultivos alternativos con menor HH y PMR comparable."""
    df_maestro = pd.read_csv(MAESTRO_CSV)

    crop_agg = (
        df_maestro.groupby("Cultivo")
        .agg({
            "HHtotal_m3_ton": "mean",
            "HHazul_m3_ton": "mean",
            "Rend_t_ha": "mean",
            "PMR": "mean",
            "VolumenTotal_t": "sum",
        })
        .reset_index()
    )
    crop_agg = crop_agg[crop_agg["VolumenTotal_t"] > 1000].copy()

    trigo_row = crop_agg[crop_agg["Cultivo"] == "Trigo grano"].iloc[0]
    trigo_hh = trigo_row["HHtotal_m3_ton"]
    trigo_pmr = trigo_row["PMR"]

    # Cultivos con menor HH que el trigo y más de 1000 ton de producción
    alternatives = crop_agg[crop_agg["HHtotal_m3_ton"] < trigo_hh].sort_values("HHtotal_m3_ton").head(10)
    # Add trigo at the end
    plot_data = pd.concat([alternatives, crop_agg[crop_agg["Cultivo"] == "Trigo grano"]], ignore_index=True)
    plot_data = plot_data.sort_values("HHtotal_m3_ton", ascending=True)

    fig, ax = plt.subplots(figsize=(13, 7), dpi=300)
    x_pos = range(len(plot_data))
    colors = ["#ef4444" if c == "Trigo grano" else "#3b82f6" for c in plot_data["Cultivo"]]
    bars = ax.bar(x_pos, plot_data["HHtotal_m3_ton"], color=colors, alpha=0.9, edgecolor="#333", linewidth=0.5)

    for i, (_, row) in enumerate(plot_data.iterrows()):
        ahorro = trigo_hh - row["HHtotal_m3_ton"]
        pct = ahorro / trigo_hh * 100
        if row["Cultivo"] != "Trigo grano":
            ax.annotate(
                f"-{pct:.0f}%", (i, row["HHtotal_m3_ton"]),
                textcoords="offset points", xytext=(0, 5), ha="center",
                fontsize=8, color="#1e293b", fontweight="bold",
            )

    ax.set_xticks(x_pos)
    ax.set_xticklabels(plot_data["Cultivo"], rotation=45, ha="right", fontsize=9)
    ax.set_ylabel("Huella Hídrica Total (m³/ton)")
    ax.set_title(
        f"Comparativa: Cultivos Alternativos vs. Trigo Grano (HH = {trigo_hh:.0f} m³/ton)",
        pad=15, weight="bold",
    )
    ax.axhline(y=trigo_hh, color="#ef4444", linestyle="--", alpha=0.6, label=f"Trigo grano ({trigo_hh:.0f} m³/ton)")

    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor="#3b82f6", edgecolor="black", label="Cultivo alternativo"),
        Patch(facecolor="#ef4444", edgecolor="black", label="Trigo grano (referencia)"),
    ]
    ax.legend(handles=legend_elements, loc="upper left", frameon=True, shadow=True)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "trigo_vs_alternativas.png", bbox_inches="tight")
    plt.close()

    print("Comparativa trigo vs alternativas generada exitosamente.")


def plot_trigo_variabilidad_municipal():
    """Genera gráfica de variabilidad del HH del trigo grano en los top 10 municipios productores."""
    df_maestro = pd.read_csv(MAESTRO_CSV)
    trigo = df_maestro[df_maestro["Cultivo"] == "Trigo grano"].copy()

    top_mun = (
        trigo.groupby("Municipio")
        .agg({"VolumenTotal_t": "sum", "HHtotal_m3_ton": "mean", "HHazul_m3_ton": "mean", "HHverde_m3_ton": "mean"})
        .sort_values("VolumenTotal_t", ascending=False)
        .head(10)
        .sort_values("HHtotal_m3_ton", ascending=True)
    )

    fig, ax = plt.subplots(figsize=(12, 7), dpi=300)
    y_pos = range(len(top_mun))
    ax.barh(y_pos, top_mun["HHazul_m3_ton"].values, color="#38bdf8", label="HH Azul", alpha=0.9)
    ax.barh(
        y_pos, top_mun["HHverde_m3_ton"].values,
        left=top_mun["HHazul_m3_ton"].values, color="#4ade80", label="HH Verde", alpha=0.9,
    )
    ax.set_yticks(y_pos)
    ax.set_yticklabels(top_mun.index, fontsize=10)
    ax.set_xlabel("Huella Hídrica Total (m³/ton)")
    ax.set_title("Huella Hídrica del Trigo Grano en los 10 Principales Municipios Productores", pad=15, weight="bold")
    ax.legend(loc="lower right", frameon=True, shadow=True)

    for i, (mun, row) in enumerate(top_mun.iterrows()):
        ax.annotate(
            f'{row["HHtotal_m3_ton"]:.0f}', (row["HHtotal_m3_ton"], i),
            textcoords="offset points", xytext=(5, 0), ha="left", fontsize=8,
        )

    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "trigo_hh_municipios.png", bbox_inches="tight")
    plt.close()

    print("Variabilidad del HH del trigo por municipio generada exitosamente.")


def main():
    print("Iniciando análisis espacial de sequía y vulnerabilidad...")
    gdf, df_indices, repna = load_data()
    map_municipal_isag(gdf, df_indices)
    analyze_ddr_pressure(gdf, df_indices, repna)
    generate_ddr_technification_map(gdf, df_indices)
    run_extended_analysis(gdf, df_indices)
    analyze_dam_volumes()
    compute_composite_vulnerability(gdf, df_indices)
    generate_efficiency_productivity_maps(gdf, df_indices)
    # New functions for report figures
    plot_ranking_hh()
    compute_ddr_summary_table()
    plot_cuadrante_hh_pmr()
    plot_ddr_cajeme_profile()
    plot_trigo_vs_alternatives()
    plot_trigo_variabilidad_municipal()
    print("Análisis espacial completado exitosamente.")


if __name__ == "__main__":
    main()
